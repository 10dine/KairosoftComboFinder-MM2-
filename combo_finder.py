#tkinter
import tkinter as tk

#excel 
from openpyxl import load_workbook
wb = load_workbook(filename = 'MM2.xlsx')
shtr1 = wb['Combo_list']
shtr2 = wb['Facilities_list']

#excel list 2
fclist = []
for col in shtr2['A']:
    fclist.append(col.value)

#main
class CF: #CF as in Combo Finder
    
    def __init__(self, master):
        #master
        self.master = master
        self.frame = tk.Frame(self.master)
        self.frame.configure(bg="white")
        #logo
        self.logo = tk.PhotoImage(file="img1.gif")
        self.label1 = tk.Label (image=self.logo, bg="white")
        self.label1.pack()
        #dropdown_selector
        pick = tk.StringVar()
        pick.set(fclist[0]) # default value
        self.w = tk.OptionMenu(master, pick, *fclist)
        self.w.pack()
        #gobutton
        self.button1 = tk.Button(self.frame, text = 'Get Combos', width = 25, bg="white", command=lambda : self.combo_finder1(pick.get())) # lambda makes sure that command goes after click
        self.button1.pack()

        #clearlistbutton | did not work will try later one day
        #self.button2 = tk.Button(self.frame, text = 'Clear combo list!', width = 25, bg="white", command = lambda : self.clearlist).pack()

        #cmlabel
        self.cmlabel = tk.Label(text = "-, -, -\n-, -, -\n-, -, -")
        self.cmlabel.pack()
        #packs window
        self.frame.pack()

    #Skims through excel shtr 1 and replace text in cmlabel with combos  
    def combo_finder1 (self, chosen):
        cl=['N/a','N/a','N/a','N/a','N/a','N/a','N/a','N/a']
        #main for loop
        for row in shtr1.iter_rows(1, 44): #iterates through each rows
            for cell in row:               #iterates through each cell from given row
                if cell.value == chosen:   #checks if given arg is equal value in cell
                    for x in row[1:4]:     #iterates through each row that contained arg
                        cl.insert(0, x.value) #inserts all values of cells from the row that contained arg tothe front of 'cl' list.
        #label text and modifier
        labeltext = tk.StringVar() #used string as I was aticipating an updating label but forgot it was a function
        labeltext = "{}, {}, {}\n{}, {}, {}\n{}, {}, {}".format(cl[0], cl[1], cl[2], cl[3], cl[4], cl[5], cl[6], cl[7], cl[8])                        
        self.cmlabel.config(text = labeltext)

#    def clearlist(self): | Did not work will try again later
#        cl=['N/a','N/a','N/a','N/a','N/a','N/a','N/a','N/a']
#        self.cmlabel.pack()

#main def that initiates(?) tkinter and runs class CF
def main(): 
    root = tk.Tk()
    app = CF(root)
    root.mainloop()

#initiates the code
if __name__ == '__main__':
    main()